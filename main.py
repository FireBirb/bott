import sys
import vk_api
from vk_api import VkUpload
from vk_api.longpoll import VkLongPoll, VkEventType
from vk_api.utils import get_random_id
from vk_api.keyboard import VkKeyboard, VkKeyboardColor
import requests
from bs4 import BeautifulSoup
import openpyxl
import re
import json
import PIL.Image as Image
import matplotlib.pyplot as plt
import numpy as np
import datetime
import calendar
import time
from datetime import date


def main():
    cal = calendar.Calendar()
    num_week = -1
    year = date.today().year
    for x in range(3, date.today().month + 1):
        for y in (cal.monthdatescalendar(year=year, month=x)):
            num_week += 1
    for k in cal.itermonthdays2(year=year, month=date.today().month):
        if k[0] == date.today().day:
            day_of_week = k[1] + 1

    month = {
        1:"января",
        2:"феврля",
        3:"марта",
        4:"апреля",
        5:"мая",
        6:"июня",
        7:"июля",
        8:"августа",
        9:"сентября",
        10:"октября",
        11:"ноября",
        12:"декабря"
    }
    week = {
        1: "понедельник",
        2: "вторник",
        3: "среда",
        4: "четверг",
        5: "пятница",
        6: "суббота",
        7: "воскресенье"
    }

    page = requests.get("https://www.mirea.ru/schedule/")
    soup = BeautifulSoup(page.text, "html.parser")

    links = soup.find("div", {"class": "rasspisanie"}). \
        find(string="Институт информационных технологий"). \
        find_parent("div"). \
        find_parent("div"). \
        findAll("a", {"class": "uk-link-toggle"})
    link = []

    # page = requests.get('https://coronavirusstat.ru/country/moskva/')
    # soup = BeautifulSoup(page.text, "html.parser")
    # c_result = soup.findAll("table")[0].find("tbody").findAll("tr")
    # date = []
    # num = []
    # for x in c_result:
    #    date += x.findAll("th")[0]
    # for y in c_result:
    #    num += x.findAll("td")[0]
    # print (num)

    # fig, ax = plt.subplots()
    # ax.plot(date, data=num)
    # fig.savefig('covid.png')

    vk_session = vk_api.VkApi(
        token="8134344c015a55388c824d80bd8d57e94e1ecdbee616238e806d239277e65adcb50639fcbcea330550e40")
    vk = vk_session.get_api()
    longpoll = VkLongPoll(vk_session)

    # upload = VkUpload(vk_session)
    # attachments = []
    # image = requests.get("https://www.gastronom.ru/binfiles/images/20150312/bfc37d95.jpg", stream=True)
    # photo = upload.photo_messages(photos=image.raw)[0]
    # attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))

    st_keyb = VkKeyboard(one_time=True)
    st_keyb.add_button('начать', color=VkKeyboardColor.POSITIVE)

    primary_keyb = VkKeyboard(one_time=True)
    primary_keyb.add_button('на сегодня', color=VkKeyboardColor.POSITIVE)
    primary_keyb.add_button('на завтра', color=VkKeyboardColor.NEGATIVE)
    primary_keyb.add_line()
    primary_keyb.add_button('на эту неделю', color=VkKeyboardColor.PRIMARY)
    primary_keyb.add_button('на следующую неделю', color=VkKeyboardColor.PRIMARY)
    primary_keyb.add_line()
    primary_keyb.add_button('какая неделя?', color=VkKeyboardColor.SECONDARY)
    primary_keyb.add_button('какая группа?', color=VkKeyboardColor.SECONDARY)

    groups = openpyxl.load_workbook("groups.xlsx")
    gr_sheet = groups.active
    # print (gr_sheet.max_row)
    num_cols = 2
    num_rows = gr_sheet.max_row

    ids = {}
    for i in range(1, num_rows + 1):
        if gr_sheet[i][0].value:
            ids |= {(gr_sheet[i][0].value): i}
    print(ids)

    # week =

    vk.messages.send(
        user_id=247136183,
        random_id=get_random_id(),
        message='Бот запущен'
    )

    while True:
        for event in longpoll.listen():
            if event.type == VkEventType.MESSAGE_NEW and event.text == "Привет! Начнем?":
                vk.messages.send(
                    user_id=event.user_id,
                    random_id=get_random_id(),
                    keyboard=st_keyb.get_keyboard()
                )
            if event.type == VkEventType.MESSAGE_NEW and event.to_me and event.text:
                text = event.text.lower()
                print('New from {}, text = {}'.format(event.user_id, event.text))
                if text == 'умри' and event.user_id == 247136183:
                    sys.exit()
                elif text == 'начать':
                    vk.messages.send(
                        user_id=event.user_id,
                        random_id=get_random_id(),
                        message='!!Придумать инструкцию сюда!!'
                    )
                elif re.search("^И[А-Я]БО-\d+-\d+$", text, re.IGNORECASE):
                    st = re.findall("\d+$", text)
                    st = round(year + (date.today().month / 12) - int(st[0]) - 2000)
                    if not (event.user_id in ids.keys()):
                        gr_sheet.append([event.user_id, text, st])
                        num_rows += 1
                        ids |= {(gr_sheet.cell(row=num_rows, column=1).value): num_rows}
                        # print (ids)
                        # print ("Im if")
                    else:
                        i = ids[event.user_id]
                        gr_sheet.cell(row=i, column=2).value = text
                        gr_sheet.cell(row=i, column=3).value = st
                        gr_sheet.cell(row=i, column=4).value = ""
                        # print (ids)
                        # print ("Im else")
                    groups.save("groups.xlsx")
                    vk.messages.send(
                        user_id=event.user_id,
                        random_id=get_random_id(),
                        message="Я запомнил, что ты из группы {}.".format(event.text.upper())
                    )
                elif text == 'бот':
                    vk.messages.send(
                        user_id=event.user_id,
                        random_id=get_random_id(),
                        message="Показать расписание...",
                        keyboard=primary_keyb.get_keyboard()
                    )

                elif text == 'на сегодня':
                    start = "Расписание на " + str(date.today().day) + " " + str(month[date.today().month]) + ":\n"
                    message = start
                    cur_num = ids[event.user_id]
                    cur_curs = str(gr_sheet.cell(row=cur_num, column=3).value)
                    cur_gr = gr_sheet.cell(row=cur_num, column=2).value
                    for x in links:
                        link += re.findall("https://.*/ИИТ_" + cur_curs + ".*\.xlsx", str(x))
                    f = open('time_table.xlsx', "wb")
                    resp = requests.get(link[0])
                    f.write(resp.content)
                    link = []

                    t_table = openpyxl.load_workbook("time_table.xlsx")
                    t_sheet = t_table.active
                    t_num_cols = t_sheet.max_column
                    t_num_rows = t_sheet.max_row
                    if not (gr_sheet.cell(row=cur_num, column=4).value):
                        for x in range(6, t_num_cols + 1):
                            if str(t_sheet.cell(row=2, column=x).value).lower() == cur_gr:
                                gr_sheet.cell(row=cur_num, column=4).value = x
                                groups.save("groups.xlsx")
                                break
                    col = gr_sheet.cell(row=cur_num, column=4).value
                    for x in range(4, 65, 12):
                        if t_sheet.cell(row=x, column=1).value == week[day_of_week].upper():
                            n = 1
                            if num_week % 2:
                                for r in range(x, x + 11, 2):
                                    message += str(n) + ")"
                                    if t_sheet.cell(row=r, column=col).value:
                                        message += str(t_sheet.cell(row=r, column=col).value)
                                        if t_sheet.cell(row=r, column=col + 1).value:
                                            message += ", " + str(t_sheet.cell(row=r, column=col + 1).value)
                                        if t_sheet.cell(row=r, column=col + 2).value:
                                            message += ", " + str(t_sheet.cell(row=r, column=col + 2).value)
                                        if t_sheet.cell(row=r, column=col + 3).value:
                                            message += ", " + str(t_sheet.cell(row=r, column=col + 3).value)
                                        if t_sheet.cell(row=r, column=col + 4).value:
                                            message += ", " + str(t_sheet.cell(row=r, column=col + 4).value)
                                        message += "\n"
                                    else:
                                        message += " –\n"
                                    n += 1
                            else:
                                for r in range(x + 1, x + 12, 2):
                                    message += str(n) + ")"
                                    if t_sheet.cell(row=r, column=col).value:
                                        message += str(t_sheet.cell(row=r, column=col).value)
                                        if t_sheet.cell(row=r, column=col + 1).value:
                                            message += ", " + str(t_sheet.cell(row=r, column=col + 1).value)
                                        if t_sheet.cell(row=r, column=col + 2).value:
                                            message += ", " + str(t_sheet.cell(row=r, column=col + 2).value)
                                        if t_sheet.cell(row=r, column=col + 3).value:
                                            message += ", " + str(t_sheet.cell(row=r, column=col + 3).value)
                                        if t_sheet.cell(row=r, column=col + 4).value:
                                            message += ", " + str(t_sheet.cell(row=r, column=col + 4).value)
                                        message += "\n"
                                    else:
                                        message += " –\n"
                                    n += 1
                        if message != start:
                            break
                    print(message)
                    vk.messages.send(
                        user_id=event.user_id,
                        random_id=get_random_id(),
                        message=message
                    )

                elif text == 'на завтра':
                    for x in cal.itermonthdays(year=year, month=date.today().month):
                        if x>0:
                            max_d = x
                    if date.today().day + 1 > max_d:
                        start = "Расписание на 1 " + str(month[date.today().month]+1) + ":\n"
                    else:
                        start = "Расписание на " + str(date.today().day+1) + " " + str(month[date.today().month]) + ":\n"
                    message = start

                    cur_num = ids[event.user_id]
                    cur_curs = str(gr_sheet.cell(row=cur_num, column=3).value)
                    cur_gr = gr_sheet.cell(row=cur_num, column=2).value
                    for x in links:
                        link += re.findall("https://.*/ИИТ_" + cur_curs + ".*\.xlsx", str(x))
                    f = open('time_table.xlsx', "wb")
                    resp = requests.get(link[0])
                    f.write(resp.content)
                    link = []

                    t_table = openpyxl.load_workbook("time_table.xlsx")
                    t_sheet = t_table.active
                    t_num_cols = t_sheet.max_column
                    t_num_rows = t_sheet.max_row
                    if not gr_sheet.cell(row=cur_num, column=4).value:
                        for x in range(6, t_num_cols + 1):
                            if str(t_sheet.cell(row=2, column=x).value).lower() == cur_gr:
                                gr_sheet.cell(row=cur_num, column=4).value = x
                                groups.save("groups.xlsx")
                                break
                    col = gr_sheet.cell(row=cur_num, column=4).value
                    change_of_week = bool(week[day_of_week] == 'воскресенье')
                    for x in range(4, 65, 12):
                        if t_sheet.cell(row=x, column=1).value == week[(day_of_week + 1) % 7].upper():
                            n = 1
                            if (num_week + change_of_week) % 2:
                                for r in range(x, x + 11, 2):
                                    message += str(n) + ")"
                                    if t_sheet.cell(row=r, column=col).value:
                                        message += str(t_sheet.cell(row=r, column=col).value)
                                        if t_sheet.cell(row=r, column=col + 1).value:
                                            message += ", " + str(t_sheet.cell(row=r, column=col + 1).value)
                                        if t_sheet.cell(row=r, column=col + 2).value:
                                            message += ", " + str(t_sheet.cell(row=r, column=col + 2).value)
                                        if t_sheet.cell(row=r, column=col + 3).value:
                                            message += ", " + str(t_sheet.cell(row=r, column=col + 3).value)
                                        if t_sheet.cell(row=r, column=col + 4).value:
                                            message += ", " + str(t_sheet.cell(row=r, column=col + 4).value)
                                        message += "\n"
                                    else:
                                        message += " –\n"
                                    n += 1
                            else:
                                for r in range(x + 1, x + 12, 2):
                                    message += str(n) + ")"
                                    if t_sheet.cell(row=r, column=col).value:
                                        message += str(t_sheet.cell(row=r, column=col).value)
                                        if t_sheet.cell(row=r, column=col + 1).value:
                                            message += ", " + str(t_sheet.cell(row=r, column=col + 1).value)
                                        if t_sheet.cell(row=r, column=col + 2).value:
                                            message += ", " + str(t_sheet.cell(row=r, column=col + 2).value)
                                        if t_sheet.cell(row=r, column=col + 3).value:
                                            message += ", " + str(t_sheet.cell(row=r, column=col + 3).value)
                                        if t_sheet.cell(row=r, column=col + 4).value:
                                            message += ", " + str(t_sheet.cell(row=r, column=col + 4).value)
                                        message += "\n"
                                    else:
                                        message += " –\n"
                                    n += 1
                        if message != start:
                            break
                    print(message)
                    vk.messages.send(
                        user_id=event.user_id,
                        random_id=get_random_id(),
                        message=message
                    )

                elif text == 'на эту неделю':
                    cur_num = ids[event.user_id]
                    cur_curs = str(gr_sheet.cell(row=cur_num, column=3).value)
                    cur_gr = gr_sheet.cell(row=cur_num, column=2).value
                    for x in links:
                        link += re.findall("https://.*/ИИТ_" + cur_curs + ".*\.xlsx", str(x))
                    f = open('time_table.xlsx', "wb")
                    resp = requests.get(link[0])
                    f.write(resp.content)
                    link = []

                    t_table = openpyxl.load_workbook("time_table.xlsx")
                    t_sheet = t_table.active
                    t_num_cols = t_sheet.max_column
                    t_num_rows = t_sheet.max_row
                    if not gr_sheet.cell(row=cur_num, column=4).value:
                        for x in range(6, t_num_cols + 1):
                            if str(t_sheet.cell(row=2, column=x).value).lower() == cur_gr:
                                gr_sheet.cell(row=cur_num, column=4).value = x
                                groups.save("groups.xlsx")
                                break
                    col = gr_sheet.cell(row=cur_num, column=4).value
                    cur_date = date.today().day - day_of_week + 1
                    cur_month = date.today().month
                    last_pr_day = 0
                    if cur_date<1:
                        cur_month = date.today().month-1
                        for x in cal.itermonthdays(year=year, month=date.today().month-1):
                            if x != 0:
                                last_pr_day = x
                        cur_date += last_pr_day
                    for x in cal.itermonthdays(year=year, month=date.today().month):
                        if x != 0:
                            last_cur_day = x
                    last_day = max(last_cur_day, last_pr_day)
                    odd_week = num_week % 2
                    d = 1
                    n = 1
                    message = "Расписание на " + week[d] + " " + str(cur_date) + " " + month[cur_month] + ":\n\n"
                    for r in range(5 - odd_week, 76 - odd_week, 2):
                         message += str(n) + ")"
                         if t_sheet.cell(row=r, column=col).value:
                            message += str(t_sheet.cell(row=r, column=col).value)
                            if t_sheet.cell(row=r, column=col + 1).value:
                                message += ", " + str(t_sheet.cell(row=r, column=col + 1).value)
                            if t_sheet.cell(row=r, column=col + 2).value:
                                message += ", " + str(t_sheet.cell(row=r, column=col + 2).value)
                            if t_sheet.cell(row=r, column=col + 3).value:
                                message += ", " + str(t_sheet.cell(row=r, column=col + 3).value)
                            if t_sheet.cell(row=r, column=col + 4).value:
                                message += ", " + str(t_sheet.cell(row=r, column=col + 4).value)
                            message += "\n"
                         else:
                             message += " –\n"
                         n += 1
                         if n == 7:
                             n = 1
                             d += 1
                             cur_date += 1
                             if d != 7:
                                 message += "\nРасписание на " + week[d]
                                 if d == 3 or d == 5 or d == 6:
                                     message = message[:-1]
                                     message += "у"
                                 if cur_date>last_day:
                                     cur_month += 1
                                     cur_date = 1
                                 message += " " + str(cur_date) + " " + month[cur_month] + ":"
                             message += "\n\n"


                    print(message)
                    #print(message)

                    vk.messages.send(
                        user_id=event.user_id,
                        random_id=get_random_id(),
                        message=message
                    )
                elif text == 'на следующую неделю':

                    vk.messages.send(
                        user_id=event.user_id,
                        random_id=get_random_id(),
                        message="idk"
                    )
                elif text == 'какая неделя?':
                    vk.messages.send(
                        user_id=event.user_id,
                        random_id=get_random_id(),
                        message="Идет {} неделя".format(num_week)
                    )
                elif text == 'какая группа?':
                    r = ids[event.user_id]
                    message = "Показываю расписание группы {}". \
                        format(gr_sheet.cell(row=r, column=2).value.upper())
                    vk.messages.send(
                        user_id=event.user_id,
                        random_id=get_random_id(),
                        message=message
                    )
                else:
                    vk.messages.send(
                        user_id=event.user_id,
                        random_id=get_random_id(),
                        message="Неизвестная команда"
                    )


if __name__ == '__main__':
    main()
